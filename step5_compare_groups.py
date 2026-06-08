# -*- coding: utf-8 -*-



import os
import numpy as np
import pandas as pd

from typing import Optional, Dict, List, Tuple

from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import (
    roc_auc_score, roc_curve, auc,
    precision_recall_curve, average_precision_score,
    confusion_matrix, ConfusionMatrixDisplay
)
from sklearn.calibration import calibration_curve

import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage


# -----------------------------
# Config
# -----------------------------
XLSX_PATH = "data/DATA.xlsx"
OUTDIR = "compare_groups_results"
os.makedirs(OUTDIR, exist_ok=True)

FIGDIR = os.path.join(OUTDIR, "figures_final")
os.makedirs(FIGDIR, exist_ok=True)

OUT_XLSX = os.path.join(OUTDIR, "final_two_models_cv50.xlsx")
OUT_CSV  = os.path.join(OUTDIR, "final_two_models_summary.csv")

LABEL_COL = "efficacy"

N_SPLITS = 3
N_REPEATS = 50
BASE_RANDOM_STATE = 42

# Group column and derived feature
GROUP_COL = "合并用药组别(1=化疗+GD2+NK，2=GD2联合NK)"
GROUP_FEAT = "group01"  # (raw-1) => 0/1

# Final features (same structure for both subgroup models)
FINAL_FEATURES = [
    "dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "Na+_mean", "SAA_mean"
]


# -----------------------------
# Helpers
# -----------------------------
def auto_pick_sheet(xlsx_path: str) -> str:
    xls = pd.ExcelFile(xlsx_path)
    sheets = xls.sheet_names

    best_sheet, best_score, best_missing = None, -1, None

    # for auto pick, only check raw columns + label + group col
    needed = [LABEL_COL, GROUP_COL] + FINAL_FEATURES

    for s in sheets:
        try:
            df = pd.read_excel(xlsx_path, sheet_name=s, nrows=5)
            df.columns = [str(c).strip() for c in df.columns]
            cols = set(df.columns)

            missing = [c for c in needed if c not in cols]

            score = 0
            if LABEL_COL in cols:
                score += 100
            if GROUP_COL in cols:
                score += 50
            for f in FINAL_FEATURES:
                if f in cols:
                    score += 2
            for hint in ["id", "patient_id", "name", "Name"]:
                if hint in cols:
                    score += 1

            if score > best_score:
                best_score, best_sheet, best_missing = score, s, missing
        except Exception:
            continue

    if best_sheet is None:
        raise RuntimeError("[ERROR] No readable sheet in workbook.")

    print(f"[OK] Auto-picked sheet: {best_sheet} (score={best_score})")
    if best_missing:
        print(f"[INFO] Missing columns: {best_missing}")
    return best_sheet


def make_pipeline(best_params: Optional[Dict] = None):
    C = 1.0
    penalty = "l2"
    if best_params:
        C = float(best_params.get("C", C))
        penalty = str(best_params.get("penalty", penalty))

    return Pipeline([
        ("scaler", StandardScaler()),
        ("clf", LogisticRegression(solver="liblinear", max_iter=2000, C=C, penalty=penalty))
    ])


def _to_numeric_inplace(df: pd.DataFrame, cols: List[str]):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")


def add_group01(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if GROUP_COL in df.columns:
        df[GROUP_FEAT] = pd.to_numeric(df[GROUP_COL], errors="coerce") - 1.0
    else:
        df[GROUP_FEAT] = np.nan
    return df


def prepare_subset_for_model(df: pd.DataFrame, features: List[str], model_name: str) -> pd.DataFrame:
    use_cols = [LABEL_COL] + features
    missing = [c for c in use_cols if c not in df.columns]
    if missing:
        raise KeyError(f"[ERROR] Model '{model_name}' missing columns: {missing}")

    sub = df[use_cols].copy()
    sub.columns = [str(c).strip() for c in sub.columns]

    _to_numeric_inplace(sub, use_cols)

    before = len(sub)
    sub = sub.dropna()
    after = len(sub)

    sub[LABEL_COL] = sub[LABEL_COL].astype(int)
    if not set(sub[LABEL_COL].unique()).issubset({0, 1}):
        raise ValueError(f"[ERROR] {LABEL_COL} must be 0/1. Got: {sub[LABEL_COL].unique()}")

    ycnt = np.bincount(sub[LABEL_COL].values, minlength=2)
    print(f"[INFO] Model '{model_name}' subset: {after} rows (dropped {before-after}); y0={ycnt[0]} y1={ycnt[1]}")
    print(f"[INFO] Features used ({len(features)}): {features}")

    return sub


def youden_threshold(y_true, y_prob):
    fpr, tpr, thr = roc_curve(y_true, y_prob)
    j = tpr - fpr
    idx = int(np.nanargmax(j))
    return float(thr[idx])


def run_cv50(df_subset: pd.DataFrame, features: List[str], model_name: str, best_params: Optional[Dict] = None):

    X = df_subset[features].values
    y = df_subset[LABEL_COL].values

    pipe = make_pipeline(best_params=best_params)

    fold_rows, repeat_rows = [], []
    oof_rows = []
    coef_rows = []

    for rep in range(1, N_REPEATS + 1):
        seed = BASE_RANDOM_STATE + rep
        cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=seed)

        rep_aucs = []
        for fold, (tr, te) in enumerate(cv.split(X, y), start=1):
            pipe.fit(X[tr], y[tr])

            yte = y[te]
            prob = pipe.predict_proba(X[te])[:, 1]

            if len(np.unique(yte)) < 2:
                fold_auc = np.nan
            else:
                fold_auc = roc_auc_score(yte, prob)

            rep_aucs.append(fold_auc)
            fold_rows.append({
                "model": model_name,
                "repeat": rep,
                "random_state": seed,
                "fold": fold,
                "auc": fold_auc,
                "test_n": int(len(te)),
                "test_n0": int((yte == 0).sum()),
                "test_n1": int((yte == 1).sum()),
            })

            for yy, pp in zip(yte, prob):
                oof_rows.append({
                    "model": model_name,
                    "repeat": rep,
                    "fold": fold,
                    "y_true": int(yy),
                    "y_prob": float(pp),
                })

        rep_aucs = np.array(rep_aucs, dtype=float)
        repeat_rows.append({
            "model": model_name,
            "repeat": rep,
            "random_state": seed,
            "auc_mean": float(np.nanmean(rep_aucs)),
            "auc_sd": float(np.nanstd(rep_aucs)),
            "n_folds_valid_auc": int(np.isfinite(rep_aucs).sum()),
        })

        # coefficient stability
        pipe.fit(X, y)
        clf = pipe.named_steps["clf"]
        coefs = clf.coef_.ravel()
        for fname, cval in zip(features, coefs):
            coef_rows.append({
                "model": model_name,
                "repeat": rep,
                "feature": fname,
                "coef": float(cval),
            })

    fold_df = pd.DataFrame(fold_rows)
    rep_df = pd.DataFrame(repeat_rows)
    oof_df = pd.DataFrame(oof_rows)
    coef_df = pd.DataFrame(coef_rows)

    valid_fold = fold_df["auc"].dropna().values.astype(float)

    def q(x, p):
        x = np.asarray(x, dtype=float)
        x = x[np.isfinite(x)]
        return float(np.quantile(x, p)) if len(x) else np.nan

    stats = {
        "model": model_name,
        "n_samples": int(len(df_subset)),
        "n_features": int(len(features)),
        "features": ", ".join(features),

        "auc_repeat_mean": float(rep_df["auc_mean"].mean()),
        "auc_repeat_sd": float(rep_df["auc_mean"].std()),
        "auc_repeat_p10": q(rep_df["auc_mean"].values, 0.10),
        "auc_repeat_p50": q(rep_df["auc_mean"].values, 0.50),
        "auc_repeat_p90": q(rep_df["auc_mean"].values, 0.90),

        "auc_fold_mean": float(np.mean(valid_fold)) if len(valid_fold) else np.nan,
        "auc_fold_sd": float(np.std(valid_fold)) if len(valid_fold) else np.nan,
        "auc_fold_p10": q(valid_fold, 0.10),
        "auc_fold_p50": q(valid_fold, 0.50),
        "auc_fold_p90": q(valid_fold, 0.90),

        "n_folds_total": int(len(fold_df)),
        "n_folds_valid_auc": int(np.isfinite(fold_df["auc"]).sum()),
        "n_splits": int(N_SPLITS),
        "n_repeats": int(N_REPEATS),
        "base_random_state": int(BASE_RANDOM_STATE),
    }

    return fold_df, rep_df, oof_df, coef_df, stats


# -----------------------------
# Visualization helpers
# -----------------------------
def savefig(path):
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()


def plot_auc_violin(rep_all: pd.DataFrame, out_png: str):
    models = list(rep_all["model"].unique())
    data = [rep_all.loc[rep_all["model"] == m, "auc_mean"].dropna().values for m in models]

    plt.figure(figsize=(8.5, 4.8))
    plt.violinplot(data, showmeans=True, showextrema=True, showmedians=True)
    plt.xticks(np.arange(1, len(models) + 1), models, rotation=20, ha="right")
    plt.ylabel("AUC (repeat mean)")
    plt.title("FINAL models: AUC distribution across repeats (CV50)")
    savefig(out_png)


def plot_repeat_trace(rep_all: pd.DataFrame, out_png: str):
    plt.figure(figsize=(9.5, 4.8))
    for m in list(rep_all["model"].unique()):
        d = rep_all[rep_all["model"] == m].sort_values("repeat")
        plt.plot(d["repeat"].values, d["auc_mean"].values, label=m, linewidth=1.5)
    plt.xlabel("Repeat")
    plt.ylabel("AUC (repeat mean)")
    plt.title("FINAL models: repeat-wise AUC trace")
    plt.legend()
    savefig(out_png)


def plot_roc_pr_from_oof(oof_all: pd.DataFrame, out_roc_png: str, out_pr_png: str):
    models = list(oof_all["model"].unique())

    # ROC
    plt.figure(figsize=(6.5, 6))
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        fpr, tpr, _ = roc_curve(y, p)
        roc_auc = auc(fpr, tpr)
        plt.plot(fpr, tpr, label=f"{m} (AUC={roc_auc:.3f})")
    plt.plot([0, 1], [0, 1], linestyle="--")
    plt.xlabel("FPR")
    plt.ylabel("TPR")
    plt.title("FINAL models: aggregated OOF ROC")
    plt.legend()
    savefig(out_roc_png)

    # PR
    plt.figure(figsize=(6.5, 6))
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        prec, rec, _ = precision_recall_curve(y, p)
        ap = average_precision_score(y, p)
        plt.plot(rec, prec, label=f"{m} (AP={ap:.3f})")
    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.title("FINAL models: aggregated OOF PR")
    plt.legend()
    savefig(out_pr_png)


def plot_calibration(oof_all: pd.DataFrame, out_png: str):
    plt.figure(figsize=(6.5, 6))
    models = list(oof_all["model"].unique())
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        frac_pos, mean_pred = calibration_curve(y, p, n_bins=10, strategy="quantile")
        plt.plot(mean_pred, frac_pos, marker="o", linewidth=1.5, label=m)
    plt.plot([0, 1], [0, 1], linestyle="--")
    plt.xlabel("Mean predicted probability")
    plt.ylabel("Fraction of positives")
    plt.title("FINAL models: calibration (aggregated OOF)")
    plt.legend()
    savefig(out_png)


def plot_confusion_matrices(oof_all: pd.DataFrame, out_png_05: str, out_png_youden: str):
    models = list(oof_all["model"].unique())

    # 0.5 threshold
    plt.figure(figsize=(10, 4.5))
    for i, m in enumerate(models, start=1):
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        yhat = (p >= 0.5).astype(int)
        cm = confusion_matrix(y, yhat, labels=[0, 1])
        ax = plt.subplot(1, len(models), i)
        disp = ConfusionMatrixDisplay(cm, display_labels=[0, 1])
        disp.plot(ax=ax, values_format="d", colorbar=False)
        ax.set_title(f"{m}\nthr=0.5")
    savefig(out_png_05)

    # Youden threshold
    plt.figure(figsize=(10, 4.5))
    for i, m in enumerate(models, start=1):
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        thr = youden_threshold(y, p)
        yhat = (p >= thr).astype(int)
        cm = confusion_matrix(y, yhat, labels=[0, 1])
        ax = plt.subplot(1, len(models), i)
        disp = ConfusionMatrixDisplay(cm, display_labels=[0, 1])
        disp.plot(ax=ax, values_format="d", colorbar=False)
        ax.set_title(f"{m}\nYouden thr={thr:.2f}")
    savefig(out_png_youden)


def plot_coef_stability(coef_df: pd.DataFrame, out_png: str):
    models = list(coef_df["model"].unique())
    n = len(models)
    plt.figure(figsize=(11, 4.2 * max(1, n)))

    for idx, m in enumerate(models, start=1):
        d = coef_df[coef_df["model"] == m]
        g = d.groupby("feature")["coef"].agg(["mean", "std"]).reset_index()
        g["abs_mean"] = g["mean"].abs()
        g = g.sort_values("abs_mean", ascending=False)

        ax = plt.subplot(n, 1, idx)
        ax.errorbar(
            x=np.arange(len(g)),
            y=g["mean"].values,
            yerr=g["std"].values,
            fmt="o",
            capsize=3
        )
        ax.axhline(0, linestyle="--", linewidth=1)
        ax.set_xticks(np.arange(len(g)))
        ax.set_xticklabels(g["feature"].values, rotation=30, ha="right")
        ax.set_ylabel("coef (standardized)")
        ax.set_title(f"Coefficient stability: {m} (mean ± sd)")
    savefig(out_png)


def add_images_to_excel(xlsx_path: str, image_paths: List[Tuple[str, str]]):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    for sheet_name, png in image_paths:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        ws["A1"] = f"Figure: {os.path.basename(png)}"
        try:
            img = XLImage(png)
            img.anchor = "A3"
            ws.add_image(img)
        except Exception as e:
            ws["A2"] = f"[WARN] failed to embed image: {e}"
    wb.save(xlsx_path)


def robust_save_excel(excel_path: str, writer_fn):
    """
    Write to tmp then replace to avoid occasional openpyxl save timeout / lock issues.
    """
    tmp_xlsx = excel_path + ".tmp.xlsx"
    try:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
    except Exception:
        pass

    writer_fn(tmp_xlsx)
    os.replace(tmp_xlsx, excel_path)


# -----------------------------
# Main
# -----------------------------
def main():
    sheet = auto_pick_sheet(XLSX_PATH)
    df = pd.read_excel(XLSX_PATH, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]

    # add group01
    df = add_group01(df)
    _to_numeric_inplace(df, [GROUP_FEAT] + FINAL_FEATURES + [LABEL_COL])

    # split into two subsets
    df_g0 = df[df[GROUP_FEAT] == 0].copy()
    df_g1 = df[df[GROUP_FEAT] == 1].copy()

    final_models = [
        ("FINAL_group0", df_g0),
        ("FINAL_group1", df_g1),
    ]

    summary_rows = []
    excel_sheets = {"sheet_used_info": pd.DataFrame([{"sheet_used": sheet}])}

    rep_all_list, fold_all_list, oof_all_list, coef_all_list = [], [], [], []

    # run two final models
    for model_name, dfx in final_models:
        print("\n" + "="*80)
        print(f"[RUN] {model_name}")

        df_sub = prepare_subset_for_model(dfx, FINAL_FEATURES, model_name=model_name)

        fold_df, rep_df, oof_df, coef_df, stats = run_cv50(df_sub, FINAL_FEATURES, model_name)

        summary_rows.append(stats)
        rep_all_list.append(rep_df)
        fold_all_list.append(fold_df)
        oof_all_list.append(oof_df)
        coef_all_list.append(coef_df)

        excel_sheets[f"{model_name}_subset"] = df_sub
        excel_sheets[f"{model_name}_auc_per_fold"] = fold_df
        excel_sheets[f"{model_name}_auc_by_repeat"] = rep_df
        excel_sheets[f"{model_name}_oof_preds"] = oof_df
        excel_sheets[f"{model_name}_coef_by_repeat"] = coef_df
        excel_sheets[f"{model_name}_stats"] = pd.DataFrame([stats])

    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print("\n[Saved CSV]")
    print(" -", OUT_CSV)

    rep_all = pd.concat(rep_all_list, ignore_index=True) if rep_all_list else pd.DataFrame()
    oof_all = pd.concat(oof_all_list, ignore_index=True) if oof_all_list else pd.DataFrame()
    coef_all = pd.concat(coef_all_list, ignore_index=True) if coef_all_list else pd.DataFrame()

    # write excel (robust)
    def _write_excel(path):
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            summary_df.to_excel(w, sheet_name="final_summary", index=False)
            for name, dfx in excel_sheets.items():
                dfx.to_excel(w, sheet_name=name[:31], index=False)

    robust_save_excel(OUT_XLSX, _write_excel)
    print("\n[Saved Excel]")
    print(" -", OUT_XLSX)

    # figures
    fig_paths = []

    if len(rep_all):
        p1 = os.path.join(FIGDIR, "final_compare_auc_violin.png")
        plot_auc_violin(rep_all, p1)
        fig_paths.append(("fig_auc_violin", p1))

        p2 = os.path.join(FIGDIR, "final_compare_repeat_trace.png")
        plot_repeat_trace(rep_all, p2)
        fig_paths.append(("fig_repeat_trace", p2))

    if len(oof_all):
        p3 = os.path.join(FIGDIR, "final_compare_oof_roc.png")
        p4 = os.path.join(FIGDIR, "final_compare_oof_pr.png")
        plot_roc_pr_from_oof(oof_all, p3, p4)
        fig_paths.append(("fig_oof_roc", p3))
        fig_paths.append(("fig_oof_pr", p4))

        p5 = os.path.join(FIGDIR, "final_compare_calibration.png")
        plot_calibration(oof_all, p5)
        fig_paths.append(("fig_calibration", p5))

        p6 = os.path.join(FIGDIR, "final_cm_thr05.png")
        p7 = os.path.join(FIGDIR, "final_cm_youden.png")
        plot_confusion_matrices(oof_all, p6, p7)
        fig_paths.append(("fig_cm_thr05", p6))
        fig_paths.append(("fig_cm_youden", p7))

    if len(coef_all):
        p8 = os.path.join(FIGDIR, "final_compare_coef_stability.png")
        plot_coef_stability(coef_all, p8)
        fig_paths.append(("fig_coef_stability", p8))

    # embed figs
    if fig_paths:
        add_images_to_excel(OUT_XLSX, fig_paths)

    print("\n[Saved Figures]")
    print(" -", FIGDIR)
    print("\nDONE.")


if __name__ == "__main__":
    main()
