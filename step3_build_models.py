# -*- coding: utf-8 -*-



import os
import warnings
import numpy as np
import pandas as pd

from typing import Optional, Dict, List, Tuple

from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import (
    roc_auc_score, roc_curve, auc,
    precision_recall_curve, average_precision_score,
    confusion_matrix, ConfusionMatrixDisplay
)
from sklearn.calibration import calibration_curve

import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage

from scipy.stats import wilcoxon


# -----------------------------
# Silence repeated sklearn FutureWarning spam
# (keeps behavior unchanged; just stops console flooding)
# -----------------------------
warnings.filterwarnings(
    "ignore",
    category=FutureWarning,
    message=r".*'penalty' was deprecated in version 1\.8.*"
)


# -----------------------------
# Config
# -----------------------------
XLSX_PATH = "data/DATA.xlsx"
OUTDIR = "build_models_results"
os.makedirs(OUTDIR, exist_ok=True)

FIGDIR = os.path.join(OUTDIR, "figures")
os.makedirs(FIGDIR, exist_ok=True)

OUT_XLSX = os.path.join(OUTDIR, "baseline_compare_cv50.xlsx")
OUT_CSV  = os.path.join(OUTDIR, "baseline_compare_summary.csv")

LABEL_COL = "efficacy"

N_SPLITS = 3
N_REPEATS = 50
BASE_RANDOM_STATE = 42

# ---- Interaction term names (new columns we will create) ----
INT_SLOPE_RDW = "int_dyn_slope_BA_on_PCT_x_RDW-SD_mean"
INT_SAMEDIR_RDW = "int_dyn_same_dir_rate_x_RDW-SD_mean"
INT_WBC_HGB = "int_WBC_mean_x_HGB_mean"

# ---- STEPWISE MODELS (start from slope, add one-by-one) ----
# You can reorder/add steps here without changing the rest of the code.
STEPWISE_MODELS: List[Tuple[str, List[str]]] = [

    ("M0_slope",
     ["dyn_slope_BA_on_PCT"
     ]),

    ("M1_+HGB",
     ["dyn_slope_BA_on_PCT",
      "HGB_mean"
      ]),

    ("M2_+RDWSD",
     ["dyn_slope_BA_on_PCT",
      "HGB_mean",
      "RDW-SD_baseline"
      ]),

    ("M3_+same_dir",
     ["dyn_slope_BA_on_PCT",
      "HGB_mean",
      "RDW-SD_baseline",
      "dyn_same_dir_rate"
      ]),

    # ("M4_+corr",
    #  ["dyn_slope_BA_on_PCT",
    #   "HGB_mean",
    #   "RDW-SD_mean",
    #   "dyn_same_dir_rate",
    #   "dyn_corr_PCT_BA"
    #   ]),

#     ("M5_+ WBC",
#      ["dyn_slope_BA_on_PCT",
#       "WBC_mean",
#       "HGB_mean",
#       "RDW-SD_mean",
#       "dyn_same_dir_rate",
#       ]),
#
#      ("M6_+ HRR",
#      [
#       "dyn_slope_BA_on_PCT",
#       "HGB_mean",
#       "RDW-SD_mean",
#       "dyn_same_dir_rate",
#       "HRR_mean"
#       ]),
#
#     ("M7_+ CLR",
#      [
#       "dyn_slope_BA_on_PCT",
#       "HGB_mean",
#       "RDW-SD_mean",
#       "dyn_same_dir_rate",
#       "CLR_mean",
#     #   "HRR_mean"
#       ]),
#
#     ("M8_+NLR",
#      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "NLR_mean"]),
#
#     ("M9_+PLR",
#      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "PLR_mean"]),
#
#     ("M10_+LMR",
#      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "LMR_mean"]),
#
#     ("M11_+SII",
#      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "SII_mean"]),

    ("M4_+SAA",
     ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "SAA_mean"]),

    ("M5_+CHE",
     ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "SAA_mean", "CHE_slope"]),

    ("M6—m4_+CHE",
      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "CHE_slope"]),

    ("M7-m4_+Na+",
      ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "Na+_mean"]),

    ("M14_M13+SAA",
     ["dyn_slope_BA_on_PCT","HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "Na+_mean", "SAA_mean"]),

    ("M15",
     ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_baseline", "dyn_same_dir_rate", "Na+_mean", "SAA_mean", "CHE_slope"])



    # ("M16_+CAR",
    #  ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate", "CAR_mean"]),   # 效果不佳，AUC Repeat Mean = 0.75，AUC Repeat SD = 0.052


    # ("M17_+AllRatios",
    #  ["dyn_slope_BA_on_PCT", "HGB_mean", "RDW-SD_mean", "dyn_same_dir_rate",
    #   "NLR_mean", "PLR_mean", "LMR_mean", "SII_mean"])                                        # 效果不佳，AUC Repeat Mean = 0.78，AUC Repeat SD = 0.070
######################################### 以上新加入 ##############################################################################
]

MODELS = dict(STEPWISE_MODELS)

# 用于自动选 sheet 的“候选列集合”
PICK_FEATURES = sorted(list({f for _, feats in STEPWISE_MODELS for f in feats}))


# -----------------------------
# Helpers
# -----------------------------
# 自动从 Excel 中选包含「标签 + 所有候选特征」的 Sheet，无需手动指定
def auto_pick_sheet(xlsx_path: str) -> str:
    xls = pd.ExcelFile(xlsx_path)
    sheets = xls.sheet_names

    best_sheet, best_score, best_missing = None, -1, None

    for s in sheets:
        try:
            df = pd.read_excel(xlsx_path, sheet_name=s, nrows=5)
            df.columns = [str(c).strip() for c in df.columns]
            cols = set(df.columns)

            missing = []
            if LABEL_COL not in cols:
                missing.append(LABEL_COL)
            for f in PICK_FEATURES:
                if f not in cols:
                    missing.append(f)

            score = 0
            if LABEL_COL in cols:
                score += 100
            for f in PICK_FEATURES:
                if f in cols:
                    score += 2
            for hint in ["id", "patient_id", "name", "Name"]:
                if hint in cols:
                    score += 1

            if score > best_score:
                best_score, best_sheet, best_missing = score, s, missing
        except Exception:
            continue

    if best_sheet is None:
        raise RuntimeError("[ERROR] No readable sheet in workbook.")

    print(f"[OK] Auto-picked sheet: {best_sheet} (score={best_score})")
    if best_missing:
        print(f"[INFO] Missing columns (only affects some steps): {best_missing[:20]}{' ...' if len(best_missing)>20 else ''}")
    return best_sheet

# 构建「标准化 + 逻辑回归」的模型管道，兼容 sklearn 1.8 + 版本
def make_pipeline(best_params: Optional[Dict] = None):
    """
    Fixes sklearn>=1.8 FutureWarning spam:
    - If using default L2, do NOT explicitly pass penalty (keeps identical behavior).
    - For other penalties, keep backward-compatible behavior; warning is silenced globally above.
    """
    C = 1.0
    penalty = "l2"
    if best_params:
        C = float(best_params.get("C", C))
        penalty = str(best_params.get("penalty", penalty))

    lr_kwargs = dict(solver="liblinear", max_iter=2000, C=C)

    # Default behavior is L2; don't pass penalty to avoid FutureWarning in sklearn>=1.8
    if penalty != "l2":
        lr_kwargs["penalty"] = penalty  # keep original behavior for non-default options

    return Pipeline([
        ("scaler", StandardScaler()),
        ("clf", LogisticRegression(**lr_kwargs))
    ])

# 将指定列转为数值类型（原地修改），无法转换的设为 NaN
def _to_numeric_inplace(df: pd.DataFrame, cols: List[str]):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

# 生成预定义的交互特征（特征相乘），丰富特征维度
def add_all_interactions(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    base_cols = ["dyn_slope_BA_on_PCT", "dyn_same_dir_rate", "RDW-SD_mean", "WBC_mean", "HGB_mean"]
    _to_numeric_inplace(df, base_cols)

    if ("dyn_slope_BA_on_PCT" in df.columns) and ("RDW-SD_mean" in df.columns):
        df[INT_SLOPE_RDW] = df["dyn_slope_BA_on_PCT"] * df["RDW-SD_mean"]
    else:
        df[INT_SLOPE_RDW] = np.nan

    if ("dyn_same_dir_rate" in df.columns) and ("RDW-SD_mean" in df.columns):
        df[INT_SAMEDIR_RDW] = df["dyn_same_dir_rate"] * df["RDW-SD_mean"]
    else:
        df[INT_SAMEDIR_RDW] = np.nan

    if ("WBC_mean" in df.columns) and ("HGB_mean" in df.columns):
        df[INT_WBC_HGB] = df["WBC_mean"] * df["HGB_mean"]
    else:
        df[INT_WBC_HGB] = np.nan

    return df

# 为单个模型筛选、清洗数据，返回专属干净数据集
def prepare_subset_for_model(df: pd.DataFrame, features: List[str], model_name: str) -> pd.DataFrame:
    use_cols = [LABEL_COL] + features                                              # 步骤1：确定模型需要的所有列（标签列 + 特征列）
    missing = [c for c in use_cols if c not in df.columns]                         # 步骤2：检查必要列是否存在，避免列缺失导致报错
    if missing:
        raise KeyError(f"[ERROR] Model '{model_name}' missing columns: {missing}")

    sub = df[use_cols].copy()                                                      # 步骤3：筛选列并创建副本（关键：避免修改原始数据）
    sub.columns = [str(c).strip() for c in sub.columns]                            # 步骤4：清理列名（去除首尾空白字符）

    _to_numeric_inplace(sub, use_cols)                                             # 步骤5：将列转换为数值类型（原地修改）

    before = len(sub)                                                              # 步骤6：删除缺失值并统计行数变化
    sub = sub.dropna()
    after = len(sub)

    sub[LABEL_COL] = sub[LABEL_COL].astype(int)                                    # 步骤7：验证标签合法性
    if not set(sub[LABEL_COL].unique()).issubset({0, 1}):
        raise ValueError(f"[ERROR] {LABEL_COL} must be 0/1. Got: {sub[LABEL_COL].unique()}")

    ycnt = np.bincount(sub[LABEL_COL].values, minlength=2)                         # 步骤8：输出日志，监控数据分布
    print(f"[INFO] Model '{model_name}' subset: {after} rows (dropped {before-after}); y0={ycnt[0]} y1={ycnt[1]}")
    print(f"[INFO] Features used ({len(features)}): {features}")

    return sub

# 	计算 ROC 曲线的 Youden 最优阈值
def youden_threshold(y_true, y_prob):
    fpr, tpr, thr = roc_curve(y_true, y_prob)
    j = tpr - fpr
    idx = int(np.nanargmax(j))
    return float(thr[idx])


def run_cv50(df_subset: pd.DataFrame, features: List[str], model_name: str, best_params: Optional[Dict] = None):
    """
    Returns:
      fold_df: each fold AUC
      rep_df: each repeat AUC (aligned to step8: AUC computed on the concatenated OOF of that repeat)
      oof_df: out-of-fold predictions (y_true, y_prob, repeat, fold)
      coef_df: coefficients from per-repeat full fit (for stability viz)
      stats: summary dict

    IMPORTANT (alignment change)
    ----------------------------
    - step8 repeat-wise AUC is computed as: AUC(y, oof_prob_repeat) where oof_prob_repeat is the
      concatenated OOF prediction across the 3 folds within that repeat.
    - Previously step3 used mean(fold AUC) as repeat AUC; that can differ from step8.
    - This function now matches step8's definition while keeping the rest of the script unchanged.
    """
    X = df_subset[features].values
    y = df_subset[LABEL_COL].values

    pipe = make_pipeline(best_params=best_params)

    fold_rows, repeat_rows = [], []
    oof_rows = []
    coef_rows = []

    # Build ONE deterministic set of repeated stratified splits (same idea as step8)
    rskf = RepeatedStratifiedKFold(
        n_splits=N_SPLITS,
        n_repeats=N_REPEATS,
        random_state=BASE_RANDOM_STATE
    )
    splits = []
    for i, (tr, te) in enumerate(rskf.split(np.zeros_like(y), y)):
        rep = i // N_SPLITS + 1
        fold = i % N_SPLITS + 1
        splits.append((rep, fold, tr, te))

    for rep in range(1, N_REPEATS + 1):
        rep_splits = [s for s in splits if s[0] == rep]

        # Collect OOF of THIS repeat to compute repeat-wise AUC in the same way as step8
        oof_prob = np.full(len(y), np.nan, dtype=float)

        rep_fold_aucs = []
        for _, fold, tr, te in rep_splits:
            pipe.fit(X[tr], y[tr])

            yte = y[te]
            prob = pipe.predict_proba(X[te])[:, 1]

            # fold AUC (kept for diagnostics; may be NaN if test fold has single class)
            if len(np.unique(yte)) < 2:
                fold_auc = np.nan
            else:
                fold_auc = roc_auc_score(yte, prob)

            rep_fold_aucs.append(fold_auc)

            fold_rows.append({
                "model": model_name,
                "repeat": rep,
                "random_state": BASE_RANDOM_STATE,
                "fold": fold,
                "auc": fold_auc,
                "test_n": int(len(te)),
                "test_n0": int((yte == 0).sum()),
                "test_n1": int((yte == 1).sum()),
            })

            # store OOF preds (row-per-prediction; visualization only)
            for yy, pp in zip(yte, prob):
                oof_rows.append({
                    "model": model_name,
                    "repeat": rep,
                    "fold": fold,
                    "y_true": int(yy),
                    "y_prob": float(pp),
                })

            # fill repeat-level OOF vector
            oof_prob[te] = prob

        # repeat-wise AUC aligned to step8: AUC on concatenated OOF of this repeat
        mask = np.isfinite(oof_prob)
        if mask.sum() == 0 or len(np.unique(y[mask])) < 2:
            rep_auc_oof = np.nan
        else:
            rep_auc_oof = float(roc_auc_score(y[mask], oof_prob[mask]))

        rep_fold_aucs = np.asarray(rep_fold_aucs, dtype=float)
        repeat_rows.append({
            "model": model_name,
            "repeat": rep,
            "random_state": BASE_RANDOM_STATE,
            "auc_mean": rep_auc_oof,                      # <-- aligned definition
            "auc_sd": float(np.nanstd(rep_fold_aucs)),    # fold-to-fold variability within repeat (diagnostic)
            "n_folds_valid_auc": int(np.isfinite(rep_fold_aucs).sum()),
        })

        # coefficient stability: fit on full subset each repeat (kept unchanged)
        pipe.fit(X, y)
        clf = pipe.named_steps["clf"]
        coefs = clf.coef_.ravel()
        for fname, cval in zip(features, coefs):
            coef_rows.append({
                "model": model_name,
                "repeat": rep,
                "feature": fname,
                "coef": float(cval),
            })

    fold_df = pd.DataFrame(fold_rows)
    rep_df = pd.DataFrame(repeat_rows)
    oof_df = pd.DataFrame(oof_rows)
    coef_df = pd.DataFrame(coef_rows)

    valid_fold = fold_df["auc"].dropna().values.astype(float)

    def q(x, p):
        x = np.asarray(x, dtype=float)
        x = x[np.isfinite(x)]
        return float(np.quantile(x, p)) if len(x) else np.nan

    stats = {
        "model": model_name,
        "n_samples": int(len(df_subset)),
        "n_features": int(len(features)),
        "features": ", ".join(features),

        # repeat-wise AUC stats (now aligned to step8)
        "auc_repeat_mean": float(rep_df["auc_mean"].mean()),
        "auc_repeat_sd": float(rep_df["auc_mean"].std()),
        "auc_repeat_p10": q(rep_df["auc_mean"].values, 0.10),
        "auc_repeat_p50": q(rep_df["auc_mean"].values, 0.50),
        "auc_repeat_p90": q(rep_df["auc_mean"].values, 0.90),

        # fold-wise AUC stats (diagnostic)
        "auc_fold_mean": float(np.mean(valid_fold)) if len(valid_fold) else np.nan,
        "auc_fold_sd": float(np.std(valid_fold)) if len(valid_fold) else np.nan,
        "auc_fold_p10": q(valid_fold, 0.10),
        "auc_fold_p50": q(valid_fold, 0.50),
        "auc_fold_p90": q(valid_fold, 0.90),

        "n_folds_total": int(len(fold_df)),
        "n_folds_valid_auc": int(np.isfinite(fold_df["auc"]).sum()),
        "n_splits": int(N_SPLITS),
        "n_repeats": int(N_REPEATS),
        "base_random_state": int(BASE_RANDOM_STATE),
    }

    return fold_df, rep_df, oof_df, coef_df, stats
def compare_models_pvalue(rep_prev: pd.DataFrame, rep_curr: pd.DataFrame) -> float:
    """
    Paired Wilcoxon test on repeat-wise AUC means.
    Returns p-value (float) or NaN if not testable.
    """
    m = pd.merge(
        rep_prev[["repeat", "auc_mean"]],
        rep_curr[["repeat", "auc_mean"]],
        on="repeat",
        suffixes=("_prev", "_curr")
    )
    if m.empty:
        return np.nan

    a = m["auc_mean_prev"].astype(float).values
    b = m["auc_mean_curr"].astype(float).values

    # If all equal, no point testing
    if np.allclose(a, b, equal_nan=True):
        return np.nan

    # Wilcoxon requires at least one non-zero diff
    diff = b - a
    if np.allclose(diff, 0):
        return np.nan

    try:
        stat, p = wilcoxon(a, b, zero_method="wilcox", correction=False)
        return float(p)
    except Exception:
        return np.nan

def format_p_value(p, decimals=3):
    """
    Format p-value for display:
    - never show 0
    - no scientific notation
    - round to fixed decimals
    - floor at smallest shown value
    """
    if pd.isna(p):
        return ""
    p = float(p)
    min_val = 10 ** (-decimals)
    if p < min_val:
        return f"<{min_val:.{decimals}f}"
    else:
        return f"{p:.{decimals}f}"



# -----------------------------
# Visualization helpers
# -----------------------------
def savefig(path):
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()


def plot_auc_violin(rep_all: pd.DataFrame, out_png: str):
    models = list(rep_all["model"].unique())
    data = [rep_all.loc[rep_all["model"] == m, "auc_mean"].dropna().values for m in models]

    plt.figure(figsize=(10, 5))
    plt.violinplot(data, showmeans=True, showextrema=True, showmedians=True)
    plt.xticks(np.arange(1, len(models) + 1), models, rotation=20, ha="right")
    plt.ylabel("AUC (repeat mean)")
    plt.title("AUC distribution across repeats (CV50)")
    savefig(out_png)


def plot_repeat_trace(rep_all: pd.DataFrame, out_png: str):
    plt.figure(figsize=(10, 5))
    for m in list(rep_all["model"].unique()):
        d = rep_all[rep_all["model"] == m].sort_values("repeat")
        plt.plot(d["repeat"].values, d["auc_mean"].values, label=m, linewidth=1.5)
    plt.xlabel("Repeat")
    plt.ylabel("AUC (repeat mean)")
    plt.title("Repeat-wise AUC trace (stability)")
    plt.legend()
    savefig(out_png)


def plot_roc_pr_from_oof(oof_all: pd.DataFrame, out_roc_png: str, out_pr_png: str):
    """
    Aggregate OOF across all repeats+folds (per model) and draw ROC/PR.
    Note: OOF points are not independent across repeats; visualization only.
    """
    models = list(oof_all["model"].unique())

    # ROC
    plt.figure(figsize=(6.5, 6))
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        fpr, tpr, _ = roc_curve(y, p)
        roc_auc = auc(fpr, tpr)
        plt.plot(fpr, tpr, label=f"{m} (AUC={roc_auc:.3f})")
    plt.plot([0, 1], [0, 1], linestyle="--")
    plt.xlabel("FPR")
    plt.ylabel("TPR")
    plt.title("Aggregated OOF ROC (all repeats)")
    plt.legend()
    savefig(out_roc_png)

    # PR
    plt.figure(figsize=(6.5, 6))
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        prec, rec, _ = precision_recall_curve(y, p)
        ap = average_precision_score(y, p)
        plt.plot(rec, prec, label=f"{m} (AP={ap:.3f})")
    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.title("Aggregated OOF Precision-Recall (all repeats)")
    plt.legend()
    savefig(out_pr_png)


def plot_calibration(oof_all: pd.DataFrame, out_png: str):
    plt.figure(figsize=(6.5, 6))
    models = list(oof_all["model"].unique())
    for m in models:
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        frac_pos, mean_pred = calibration_curve(y, p, n_bins=10, strategy="quantile")
        plt.plot(mean_pred, frac_pos, marker="o", linewidth=1.5, label=m)
    plt.plot([0, 1], [0, 1], linestyle="--")
    plt.xlabel("Mean predicted probability")
    plt.ylabel("Fraction of positives")
    plt.title("Calibration (Aggregated OOF)")
    plt.legend()
    savefig(out_png)

def plot_calibration_single(oof_all: pd.DataFrame, model_name: str, out_png: str, n_bins: int = 10):
    """
    Draw calibration curve for ONE model only.
    Uses aggregated OOF predictions (as your existing plot does).
    """
    d = oof_all[oof_all["model"] == model_name].copy()
    if d.empty:
        raise ValueError(f"No OOF data found for model: {model_name}")

    y = d["y_true"].values.astype(int)
    p = d["y_prob"].values.astype(float)

    if len(np.unique(y)) < 2:
        raise ValueError(f"Model {model_name}: y_true has <2 classes; cannot plot calibration.")

    frac_pos, mean_pred = calibration_curve(y, p, n_bins=n_bins, strategy="quantile")

    plt.figure(figsize=(6.5, 6))
    plt.plot(mean_pred, frac_pos, marker="o", linewidth=2, label=model_name)
    plt.plot([0, 1], [0, 1], linestyle="--")
    plt.xlabel("Mean predicted probability")
    plt.ylabel("Fraction of positives")
    plt.title(f"Calibration (Aggregated OOF) — {model_name}")
    plt.legend()
    savefig(out_png)



def plot_confusion_matrices(oof_all: pd.DataFrame, out_png_05: str, out_png_youden: str):
    models = list(oof_all["model"].unique())

    # 0.5 threshold
    plt.figure(figsize=(10, 4.5))
    for i, m in enumerate(models, start=1):
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        yhat = (p >= 0.5).astype(int)
        cm = confusion_matrix(y, yhat, labels=[0, 1])
        ax = plt.subplot(1, len(models), i)
        disp = ConfusionMatrixDisplay(cm, display_labels=[0, 1])
        disp.plot(ax=ax, values_format="d", colorbar=False)
        ax.set_title(f"{m}\nthr=0.5")
    savefig(out_png_05)

    # Youden threshold (per model)
    plt.figure(figsize=(10, 4.5))
    for i, m in enumerate(models, start=1):
        d = oof_all[oof_all["model"] == m]
        y = d["y_true"].values.astype(int)
        p = d["y_prob"].values.astype(float)
        if len(np.unique(y)) < 2:
            continue
        thr = youden_threshold(y, p)
        yhat = (p >= thr).astype(int)
        cm = confusion_matrix(y, yhat, labels=[0, 1])
        ax = plt.subplot(1, len(models), i)
        disp = ConfusionMatrixDisplay(cm, display_labels=[0, 1])
        disp.plot(ax=ax, values_format="d", colorbar=False)
        ax.set_title(f"{m}\nYouden thr={thr:.2f}")
    savefig(out_png_youden)


def plot_feature_corr(df_sub: pd.DataFrame, features: List[str], out_png: str, title: str):
    X = df_sub[features].copy()
    X = X.apply(pd.to_numeric, errors="coerce")
    corr = X.corr()

    plt.figure(figsize=(7, 6))
    plt.imshow(corr.values, interpolation="nearest", aspect="auto")
    plt.xticks(range(len(features)), features, rotation=90, fontsize=8)
    plt.yticks(range(len(features)), features, fontsize=8)
    plt.title(title)
    plt.colorbar()
    savefig(out_png)


def plot_coef_stability(coef_df: pd.DataFrame, out_png: str):
    models = list(coef_df["model"].unique())
    n = len(models)
    plt.figure(figsize=(11, 4.2 * max(1, n)))

    for idx, m in enumerate(models, start=1):
        d = coef_df[coef_df["model"] == m]
        g = d.groupby("feature")["coef"].agg(["mean", "std"]).reset_index()
        g["abs_mean"] = g["mean"].abs()
        g = g.sort_values("abs_mean", ascending=False)

        ax = plt.subplot(n, 1, idx)
        ax.errorbar(
            x=np.arange(len(g)),
            y=g["mean"].values,
            yerr=g["std"].values,
            fmt="o",
            capsize=3
        )
        ax.axhline(0, linestyle="--", linewidth=1)
        ax.set_xticks(np.arange(len(g)))
        ax.set_xticklabels(g["feature"].values, rotation=30, ha="right")
        ax.set_ylabel("coef (standardized)")
        ax.set_title(f"Coefficient stability across repeats: {m} (mean ± sd)")
    savefig(out_png)



def plot_stepwise_auc_ci(rep_all: pd.DataFrame, out_png: str, step_order: list):
    """
    Mean AUC across repeats per model, with 95% CI (approx).
    """
    if rep_all.empty:
        return

    df = rep_all.copy()

    # 🔑 强制 stepwise 顺序
    df["model"] = pd.Categorical(df["model"], categories=step_order, ordered=True)
    df = df.sort_values("model")

    g = df.groupby("model")["auc_mean"].agg(["mean", "std"]).reset_index()
    g["ci95"] = 1.96 * g["std"]

    plt.figure(figsize=(10, 4.5))
    x = np.arange(len(g))
    plt.errorbar(x, g["mean"].values, yerr=g["ci95"].values, fmt="-o", capsize=4)
    plt.xticks(x, g["model"].values, rotation=30, ha="right")
    plt.ylabel("AUC (mean ± 95% CI across repeats)")
    plt.title("Stepwise model performance (CV50)")
    savefig(out_png)



def plot_increment_bar(increment_df: pd.DataFrame, out_png: str):
    """
    ΔAUC between adjacent steps + p-values (paired Wilcoxon).
    """
    if increment_df.empty:
        return

    plt.figure(figsize=(12, 5))

    x = np.arange(len(increment_df))
    y = increment_df["delta_auc_mean"].values

    bars = plt.bar(x, y)

    # 用于决定文字上下偏移量
    y_range = np.nanmax(np.abs(y))
    offset = 0.04 * y_range if y_range > 0 else 0.01

    for i, (dy, p) in enumerate(zip(y, increment_df["p_value_wilcoxon"].values)):

        if p is None or np.isnan(p):
            txt = "p=NA"
        else:
            # 统一成 3 位小数、非科学计数法
            if p < 0.0005:
                txt = "p<0.001"
            else:
                txt = f"p={p:.3f}"

        # 正柱子：标在上方；负柱子：标在下方
        if dy >= 0:
            y_text = dy + offset
            va = "bottom"
        else:
            y_text = dy - offset
            va = "top"

        plt.text(
            i,
            y_text,
            txt,
            ha="center",
            va=va,
            fontsize=10
        )

    plt.xticks(x, increment_df["to_model"].values, rotation=30, ha="right")
    plt.ylabel("Δ AUC vs previous step")
    plt.title("Incremental value of added features")

    # 让上下留点边距，避免文字被裁掉
    ymin, ymax = plt.ylim()
    plt.ylim(ymin - 0.08 * y_range, ymax + 0.12 * y_range)

    savefig(out_png)



def add_images_to_excel(xlsx_path: str, image_paths: List[Tuple[str, str]]):
    """
    image_paths: list of (sheet_name, png_path)
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    for sheet_name, png in image_paths:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        ws["A1"] = f"Figure: {os.path.basename(png)}"
        try:
            img = XLImage(png)
            img.anchor = "A3"
            ws.add_image(img)
        except Exception as e:
            ws["A2"] = f"[WARN] failed to embed image: {e}"
    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------
def main():
    sheet = auto_pick_sheet(XLSX_PATH)
    df = pd.read_excel(XLSX_PATH, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]

    # add interaction columns (even if you don't use them in stepwise)
    df = add_all_interactions(df)

    summary_rows = []
    excel_sheets = {"sheet_used_info": pd.DataFrame([{"sheet_used": sheet}])}

    rep_all_list = []
    fold_all_list = []
    oof_all_list = []
    coef_all_list = []

    corr_figs = []

    # Run stepwise models
    for model_name, feats in MODELS.items():
        print("\n" + "="*80)
        print(f"[RUN] {model_name}")

        try:
            df_sub = prepare_subset_for_model(df, feats, model_name=model_name)
        except Exception as e:
            print(f"[WARN] Skip model '{model_name}' due to subset error: {e}")
            continue

        fold_df, rep_df, oof_df, coef_df, stats = run_cv50(df_sub, feats, model_name)

        summary_rows.append(stats)

        rep_all_list.append(rep_df)
        fold_all_list.append(fold_df)
        oof_all_list.append(oof_df)
        coef_all_list.append(coef_df)

        # store per-model sheets
        excel_sheets[f"{model_name}_subset"] = df_sub
        excel_sheets[f"{model_name}_auc_per_fold"] = fold_df
        excel_sheets[f"{model_name}_auc_by_repeat"] = rep_df
        excel_sheets[f"{model_name}_oof_preds"] = oof_df
        excel_sheets[f"{model_name}_coef_by_repeat"] = coef_df
        excel_sheets[f"{model_name}_stats"] = pd.DataFrame([stats])

        # feature correlation heatmap (per model)
        corr_png = os.path.join(FIGDIR, f"{model_name}__feature_corr.png")
        plot_feature_corr(df_sub, feats, corr_png, title=f"Feature correlation: {model_name}")
        corr_figs.append((f"fig_{model_name}_corr", corr_png))

    if not summary_rows:
        raise RuntimeError("[ERROR] No model was successfully run. Check sheet/column names.")

    summary_df = pd.DataFrame(summary_rows)

    # Keep step order as in STEPWISE_MODELS (more readable than sorting by name)
    order = [m for m, _ in STEPWISE_MODELS if m in set(summary_df["model"].values)]
    summary_df["model"] = pd.Categorical(summary_df["model"], categories=order, ordered=True)
    summary_df = summary_df.sort_values("model").reset_index(drop=True)

    # Save summary CSV
    summary_df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print("\n[Saved CSV]")
    print(" -", OUT_CSV)

    # -----------------------------
    # Combine per-model dfs
    # -----------------------------
    rep_all = pd.concat(rep_all_list, ignore_index=True) if rep_all_list else pd.DataFrame()
    oof_all = pd.concat(oof_all_list, ignore_index=True) if oof_all_list else pd.DataFrame()
    coef_all = pd.concat(coef_all_list, ignore_index=True) if coef_all_list else pd.DataFrame()

    # Stepwise increment tests (paired Wilcoxon on repeat-wise AUC)
    increment_rows = []
    if len(rep_all_list) >= 2:
        for i in range(1, len(rep_all_list)):
            prev = rep_all_list[i-1]
            curr = rep_all_list[i]

            prev_model = str(prev["model"].iloc[0])
            curr_model = str(curr["model"].iloc[0])

            pval = compare_models_pvalue(prev, curr)

            delta = float(curr["auc_mean"].mean() - prev["auc_mean"].mean())
            increment_rows.append({
                "from_model": prev_model,
                "to_model": curr_model,
                "delta_auc_mean": delta,
                "p_value_wilcoxon": pval,
            })

    # increment_df = pd.DataFrame(increment_rows)
    # excel_sheets["stepwise_increment_test"] = increment_df


    # -------------------------------------------------
    # 保存▲AUC的P值
    # -------------------------------------------------
    increment_df = pd.DataFrame(increment_rows)

    # add formatted columns for display
    increment_df["delta_auc_mean_round"] = increment_df["delta_auc_mean"].round(3)
    increment_df["p_value_display"] = increment_df["p_value_wilcoxon"].apply(
        lambda x: format_p_value(x, decimals=3)
    )

    excel_sheets["stepwise_increment_test"] = increment_df


    stepwise_csv = os.path.join(OUTDIR, "step3_incremental_auc_table.csv")

    increment_df_out = increment_df.copy()
    increment_df_out["delta_auc_mean"] = increment_df_out["delta_auc_mean"].round(6)
    increment_df_out["p_value_wilcoxon"] = increment_df_out["p_value_wilcoxon"].round(6)

    increment_df_out.to_csv(stepwise_csv, index=False, encoding="utf-8-sig")
    increment_df_out["p_value_wilcoxon_fmt"] = increment_df["p_value_wilcoxon"].apply(
        lambda x: "<1e-6" if x < 1e-6 else f"{x:.3g}"
    )

    print("\n[Saved Stepwise Increment Table]")
    print(" -", stepwise_csv)
    print(increment_df_out)



    # Save one Excel (multi-sheets)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name="model_summary", index=False)
        for name, dfx in excel_sheets.items():
            safe_name = name[:31]
            dfx.to_excel(w, sheet_name=safe_name, index=False)

    print("\n[Saved Excel]")
    print(" -", OUT_XLSX)

    # -----------------------------
    # Global visualizations (compare models)
    # -----------------------------
    fig_paths = []

    if len(rep_all):
        p1 = os.path.join(FIGDIR, "compare_auc_violin.png")
        plot_auc_violin(rep_all, p1)
        fig_paths.append(("fig_compare_auc_violin", p1))

        p2 = os.path.join(FIGDIR, "compare_repeat_trace.png")
        plot_repeat_trace(rep_all, p2)
        fig_paths.append(("fig_compare_repeat_trace", p2))

        # NEW: stepwise curve (easy for non-experts)
        p_step = os.path.join(FIGDIR, "stepwise_auc_ci.png")
        plot_stepwise_auc_ci(rep_all, p_step, order)
        # plot_stepwise_auc_ci(rep_all, p_step)
        fig_paths.append(("fig_stepwise_auc_ci", p_step))

    if len(increment_df):
        p_inc = os.path.join(FIGDIR, "stepwise_increment_bar.png")
        plot_increment_bar(increment_df, p_inc)
        fig_paths.append(("fig_stepwise_increment_bar", p_inc))

    if len(oof_all):
        # OOF ROC/PR (you asked to include this)
        p3 = os.path.join(FIGDIR, "compare_oof_roc.png")
        p4 = os.path.join(FIGDIR, "compare_oof_pr.png")
        plot_roc_pr_from_oof(oof_all, p3, p4)
        fig_paths.append(("fig_compare_oof_roc", p3))
        fig_paths.append(("fig_compare_oof_pr", p4))

        p5 = os.path.join(FIGDIR, "compare_calibration.png")
        plot_calibration(oof_all, p5)
        fig_paths.append(("fig_compare_calibration", p5))

        # --- Single calibration plot for M3 ---
        # m3_name = "M3_+same_dir"
        # p5_m3 = os.path.join(FIGDIR, "calibration_M3_only.png")
        # plot_calibration_single(oof_all, m3_name, p5_m3, n_bins=10)
        # fig_paths.append(("fig_calibration_M3_only", p5_m3))
        m15_name = "M15"
        p5_m15 = os.path.join(FIGDIR, "calibration_M15_only.png")
        plot_calibration_single(oof_all, m15_name, p5_m15, n_bins=10)
        fig_paths.append(("fig_calibration_M15_only", p5_m15))


        p6 = os.path.join(FIGDIR, "compare_cm_thr05.png")
        p7 = os.path.join(FIGDIR, "compare_cm_youden.png")
        plot_confusion_matrices(oof_all, p6, p7)
        fig_paths.append(("fig_compare_cm_thr05", p6))
        fig_paths.append(("fig_compare_cm_youden", p7))

    if len(coef_all):
        p8 = os.path.join(FIGDIR, "compare_coef_stability.png")
        plot_coef_stability(coef_all, p8)
        fig_paths.append(("fig_compare_coef_stability", p8))

    # add per-model corr figs too
    fig_paths.extend(corr_figs)

    # embed key figs to Excel
    if fig_paths:
        # Put key overview figs first, then per-model corr figs
        embed_list = fig_paths[:10] + corr_figs
        add_images_to_excel(OUT_XLSX, embed_list)

    print("\n[Saved Figures]")
    print(" -", FIGDIR)

    print("\nDONE.")


if __name__ == "__main__":
    main()
